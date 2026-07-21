#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pembuat spreadsheet SIMKERMA.

Menghasilkan dua berkas:
  1. Template_Spreadsheet_MonitoringKerjasama.xlsx  (AMAN untuk repo publik — data sintetis)
     -> menyegarkan tab Pengaturan (kunci baru), mengisi kontak PIC yang kosong,
        lalu mempercantik seluruh tab.
  2. SIMKERMA_DataReal_SIAP-UPLOAD.xlsx             (LOKAL, di-gitignore — data asli)
     -> dibangun dari "DAFTAR LENGKAP REKAPAN KERJA SAMA POLTEKKES PALEMBANG (TERUPDATE).xlsx".
        Berisi nama mitra & tanggal ASLI, jadi TIDAK boleh di-commit ke repo publik.

Pemakaian:  python3 tools/build_template.py
"""
import os, re, datetime, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REAL_SRC = os.path.join(ROOT, "DAFTAR LENGKAP REKAPAN KERJA SAMA POLTEKKES PALEMBANG (TERUPDATE).xlsx")
TEMPLATE = os.path.join(ROOT, "Template_Spreadsheet_MonitoringKerjasama.xlsx")
REAL_OUT = os.path.join(ROOT, "SIMKERMA_DataReal_SIAP-UPLOAD.xlsx")
TODAY = datetime.date(2026, 7, 17)

HEADERS_MITRA = ['ID Mitra', 'Nama Mitra', 'Jenis Mitra', 'Wilayah/Provinsi',
                 'PIC Nama', 'PIC Email', 'PIC HP', 'Jumlah Kerjasama', 'Terakhir Update']
HEADERS_KERJASAMA = ['ID Kerjasama', 'Timestamp', 'ID Mitra', 'Nama Mitra', 'Jenis Mitra', 'Wilayah/Provinsi',
                     'Nomor Surat', 'Bentuk Kerja Sama', 'Ruang Lingkup', 'Pengguna MoU/PKS', 'Jabatan Penandatangan',
                     'Biaya (Rp)', 'Masa Berlaku (tahun)', 'Tanggal Mulai', 'Tanggal Berakhir',
                     'Jenis Entri', 'Ref Kerjasama Sebelumnya', 'Dokumen Induk (MoU)', 'Link File MoU/PKS', 'Catatan',
                     'Status', 'Sisa Hari', 'Diinput Oleh', 'Reminder Terakhir']

# Pengaturan — HARUS sinkron dengan _defaultSettings()/_settingKeterangan() di Code.gs
PENGATURAN = [
    ('NAMA_INSTANSI', 'Politeknik Kesehatan Kemenkes Palembang', 'Nama instansi pada notifikasi'),
    ('EMAIL_NOTIF', 'kerjasama@poltekkespalembang.ac.id', 'Email tim INTERNAL penerima rekap (pisahkan koma)'),
    ('BASE_URL', 'https://simkerma.vercel.app', 'Domain aplikasi (ubah saat pindah domain instansi)'),
    ('REMINDER_CADENCE', '90:30,60:14,30:7,7:1', 'Jadwal ingat "sisa:interval" hari. Makin dekat makin sering. Mis. 90:30,60:14,30:7,7:1'),
    ('GRACE_HABIS_HARI', 14, 'Setelah berakhir, tetap ingatkan harian sampai N hari, lalu berhenti'),
    ('EMAIL_AKTIF', 'TRUE', 'Aktifkan email rekap INTERNAL (TRUE/FALSE)'),
    ('WA_NOMOR_AKTIF', 'FALSE', 'Aktifkan WA rekap ke NOMOR perorangan tim (TRUE/FALSE)'),
    ('WA_TARGET', '', 'Nomor WA tim internal (mis. 62812xxxx), pisahkan koma'),
    ('WA_GRUP_AKTIF', 'FALSE', 'Aktifkan WA rekap ke GRUP (TRUE/FALSE)'),
    ('WA_GRUP_ID', '', 'ID grup WhatsApp (Fonnte) untuk rekap internal'),
    ('LAMPIRKAN_FILE', 'TRUE', 'Lampirkan file MoU/PKS pada email (TRUE/FALSE)'),
    ('EMAIL_EKSTERNAL_AKTIF', 'FALSE', 'Aktifkan EMAIL ke PIC mitra — hanya kerja sama miliknya (TRUE/FALSE)'),
    ('WA_EKSTERNAL_AKTIF', 'FALSE', 'HATI-HATI: WA ke PIC mitra. Bila banyak, nomor Fonnte bisa diblokir (TRUE/FALSE)'),
    ('WA_EKSTERNAL_MAKS_PER_HARI', 8, 'Batas jumlah WA eksternal per hari (anti-blokir). Mis. 8'),
    ('WA_EKSTERNAL_JEDA_DETIK', 8, 'Jeda antar kirim WA eksternal dalam detik (anti-burst). Mis. 8'),
]

# ---- Palet "Ink & Indigo" ----
INDIGO = 'FF4F46E5'
INK = 'FF1E1B4B'
GREEN = 'FF16A34A'
AMBER = 'FFD97706'
GRAY = 'FF475569'
WHITE = 'FFFFFFFF'
FILL_HABIS = PatternFill('solid', fgColor='FFFEE2E2')
FILL_SEGERA = PatternFill('solid', fgColor='FFFEF3C7')
FILL_AKTIF = PatternFill('solid', fgColor='FFDCFCE7')
FONT_HABIS = Font(color='FFB91C1C', bold=True)
FONT_SEGERA = Font(color='FF92400E', bold=True)
FONT_AKTIF = Font(color='FF166534', bold=True)
THIN = Side(style='thin', color='FFE5E7EB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def slug(s):
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    s = re.sub(r'[^a-zA-Z0-9]+', '.', s).strip('.').lower()
    return s[:40] or 'mitra'


def hitung_status(berakhir):
    if not berakhir:
        return ('Tidak Ada Tanggal', '')
    b = berakhir.date() if isinstance(berakhir, datetime.datetime) else berakhir
    sisa = (b - TODAY).days
    if sisa < 0:
        return ('Habis', sisa)
    if sisa <= 90:
        return ('Segera Berakhir', sisa)
    return ('Aktif', sisa)


def style_header(ws, ncol, tab_color):
    ws.sheet_properties.tabColor = tab_color
    fill = PatternFill('solid', fgColor=INDIGO)
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = fill
        cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{max(ws.max_row,1)}"


def color_status_column(ws, status_col_idx):
    for r in range(2, ws.max_row + 1):
        v = str(ws.cell(r, status_col_idx).value or '').strip().lower()
        cell = ws.cell(r, status_col_idx)
        if v.startswith('habis'):
            cell.fill, cell.font = FILL_HABIS, FONT_HABIS
        elif v.startswith('segera'):
            cell.fill, cell.font = FILL_SEGERA, FONT_SEGERA
        elif v.startswith('aktif'):
            cell.fill, cell.font = FILL_AKTIF, FONT_AKTIF


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def beautify_mitra(ws):
    style_header(ws, len(HEADERS_MITRA), INDIGO)
    ws.freeze_panes = 'C2'  # header + kolom ID & Nama
    set_widths(ws, [10, 40, 26, 24, 20, 30, 16, 12, 16])


def beautify_kerjasama(ws):
    style_header(ws, len(HEADERS_KERJASAMA), INK)
    ws.freeze_panes = 'E2'  # header + ID, Timestamp, ID Mitra, Nama Mitra
    set_widths(ws, [12, 18, 10, 34, 22, 22, 18, 26, 30, 30, 22,
                    14, 12, 14, 14, 12, 16, 26, 30, 24, 16, 10, 24, 18])
    color_status_column(ws, HEADERS_KERJASAMA.index('Status') + 1)


def beautify_dataset(ws):
    style_header(ws, 2, GRAY)
    ws.freeze_panes = 'A2'
    set_widths(ws, [26, 52])


def beautify_pengaturan(ws):
    style_header(ws, 3, AMBER)
    ws.freeze_panes = 'A2'
    set_widths(ws, [30, 46, 62])
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).font = Font(bold=True, color=INK)
        ws.cell(r, 3).font = Font(italic=True, color=GRAY, size=9)
        for c in (1, 2, 3):
            ws.cell(r, c).alignment = Alignment(vertical='center', wrap_text=(c == 3))


def write_pengaturan(ws):
    ws.delete_rows(1, ws.max_row)
    ws.append(['Kunci', 'Nilai', 'Keterangan'])
    for k, v, ket in PENGATURAN:
        ws.append([k, v, ket])


def ensure_pic(ws):
    """Isi PIC Email/HP yang kosong dengan placeholder JELAS-PALSU (aman)."""
    for i, r in enumerate(range(2, ws.max_row + 1)):
        nama = ws.cell(r, 2).value
        if not nama:
            continue
        if not ws.cell(r, 5).value:
            ws.cell(r, 5).value = 'Bagian Kerja Sama'
        if not ws.cell(r, 6).value:
            ws.cell(r, 6).value = f'{slug(nama)}@contoh.mitra.id'
        if not ws.cell(r, 7).value:
            ws.cell(r, 7).value = f'62812{i:07d}'


def finalize(wb, path):
    write_pengaturan(wb['Pengaturan'])
    beautify_mitra(wb['Mitra'])
    beautify_kerjasama(wb['Kerjasama'])
    beautify_dataset(wb['Dataset'])
    beautify_pengaturan(wb['Pengaturan'])
    wb.save(path)
    print(f"  -> tersimpan: {os.path.basename(path)}")


# ======================================================================
# 1) TEMPLATE AMAN (data sintetis yang sudah ada) — hanya disegarkan
# ======================================================================
def build_template():
    print("[1] Template publik (aman):")
    wb = openpyxl.load_workbook(TEMPLATE)
    ensure_pic(wb['Mitra'])
    finalize(wb, TEMPLATE)


# ======================================================================
# 2) FILE DATA ASLI (di-gitignore)
# ======================================================================
def build_real():
    if not os.path.exists(REAL_SRC):
        print("[2] Lewati file data asli — sumber tidak ditemukan.")
        return
    print("[2] File data ASLI (lokal, gitignore):")
    src = openpyxl.load_workbook(REAL_SRC, data_only=True)['Form Responses 1']
    rows = [[src.cell(r, c + 1).value for c in range(18)] for r in range(2, src.max_row + 1)]
    rows = [r for r in rows if r[5] not in (None, '')]  # hanya yang punya Nama Mitra asli
    print(f"    baris dengan nama mitra: {len(rows)}")

    mitra_idx, mitra_rows, ks_rows, count = {}, [], [], {}
    for r in rows:
        nama = str(r[5]).strip()
        jenis = str(r[4] or '').strip()
        key = nama.lower()
        if key not in mitra_idx:
            mid = f"M-{len(mitra_idx)+1:04d}"
            mitra_idx[key] = mid
            mitra_rows.append([mid, nama, jenis, '', 'Bagian Kerja Sama',
                               f'{slug(nama)}@contoh.mitra.id', f'62812{len(mitra_idx):07d}', 0, TODAY])
        mid = mitra_idx[key]
        count[mid] = count.get(mid, 0) + 1
        berakhir = r[12]
        status, sisa = hitung_status(berakhir)
        ks_rows.append([
            f"K-{len(ks_rows)+1:04d}", r[0], mid, nama, jenis, '',
            str(r[3] or '').strip(), str(r[2] or '').strip(), str(r[7] or '').strip(),
            str(r[8] or '').strip(), str(r[6] or '').strip(),
            r[9] if isinstance(r[9], (int, float)) else '', r[10] if isinstance(r[10], (int, float)) else '',
            r[11], berakhir, 'Baru', '', '', '', '',
            status, sisa, str(r[1] or '').strip(), '',
        ])
    for m in mitra_rows:
        m[7] = count.get(m[0], 0)
    print(f"    mitra unik: {len(mitra_rows)} | kerjasama: {len(ks_rows)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wm = wb.create_sheet('Mitra'); wm.append(HEADERS_MITRA)
    for m in mitra_rows: wm.append(m)
    wk = wb.create_sheet('Kerjasama'); wk.append(HEADERS_KERJASAMA)
    for k in ks_rows: wk.append(k)
    # Dataset diambil dari template supaya konsisten
    wd = wb.create_sheet('Dataset')
    for row in openpyxl.load_workbook(TEMPLATE)['Dataset'].iter_rows(values_only=True):
        wd.append(row)
    wb.create_sheet('Pengaturan')
    finalize(wb, REAL_OUT)


if __name__ == '__main__':
    build_template()
    build_real()
    print("Selesai.")
